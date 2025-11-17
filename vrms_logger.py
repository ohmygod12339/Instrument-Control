"""
Real-Time Vrms Data Logger for Keysight DSOX4034A

This script continuously reads Channel 1 Vrms from the oscilloscope every 100ms
and logs the data to Excel files with timestamps.

Features:
- Reads Vrms from Channel 1 every 100 milliseconds (precise timing)
- Records timestamp in format HH:MM:SS:MSMSMS | <Value>
- Saves to Excel file: Result_<StartDateTime>_Real-Time-Result.xlsx
- Every 5 minutes, copies data to Result_<StartDateTime>_Real-Time-Result_FINAL.xlsx
- The FINAL file can be opened for viewing without interrupting recording
- Buffered writes for fast performance

Usage:
    python vrms_logger.py <RESOURCE_STRING>

Example:
    python vrms_logger.py "USB0::0x0957::0x17A6::MY12345678::INSTR"
    python vrms_logger.py "TCPIP0::192.168.1.100::INSTR"
"""

import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Tuple
import threading
import signal

from openpyxl import Workbook, load_workbook
from instruments.keysight.dsox4034a import DSOX4034A


class VrmsLogger:
    """Real-time Vrms data logger for oscilloscope Channel 1."""

    def __init__(self, resource_string: str, results_dir: str = "results",
                 save_interval: int = 50):
        """
        Initialize the Vrms logger.

        Args:
            resource_string: VISA resource string for the oscilloscope
            results_dir: Directory to save result files (default: "results")
            save_interval: Number of measurements before saving to disk (default: 50)
        """
        self.resource_string = resource_string
        self.results_dir = Path(results_dir)
        self.save_interval = save_interval
        self.scope = None
        self.workbook = None
        self.worksheet = None
        self.main_file_path = None
        self.final_file_path = None
        self.running = False
        self.data_lock = threading.Lock()
        self.row_index = 2  # Start from row 2 (row 1 is header)
        self.start_time = None
        self.last_copy_time = None
        self.measurement_count = 0

        # Data buffer for batched writes
        self.data_buffer: List[Tuple[str, float]] = []

        # Create results directory if it doesn't exist
        self.results_dir.mkdir(exist_ok=True)

    def connect(self) -> None:
        """Connect to the oscilloscope and configure it."""
        print(f"Connecting to oscilloscope: {self.resource_string}")
        self.scope = DSOX4034A(self.resource_string)
        self.scope.connect()

        # Make sure Channel 1 is on
        self.scope.channel_on(1)

        # Set oscilloscope to RUN mode for continuous acquisition
        self.scope.run()

        print(f"Successfully connected and configured!")
        print(f"  Channel 1: ON")
        print(f"  Mode: RUN (continuous)")

    def disconnect(self) -> None:
        """Disconnect from the oscilloscope."""
        if self.scope:
            self.scope.disconnect()
            print("Disconnected from oscilloscope")

    def setup_excel_files(self) -> None:
        """Create and initialize Excel files for data logging."""
        # Generate filename with start date-time
        self.start_time = datetime.now()
        timestamp = self.start_time.strftime("%Y%m%d_%H%M%S")

        # Main recording file
        self.main_file_path = self.results_dir / f"Result_{timestamp}_Real-Time-Result.xlsx"

        # FINAL file (for viewing without interruption)
        self.final_file_path = self.results_dir / f"Result_{timestamp}_Real-Time-Result_FINAL.xlsx"

        # Create workbook and add header
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Vrms Data"

        # Add header row
        self.worksheet.append(["Timestamp", "Vrms (V)"])

        # Set column widths
        self.worksheet.column_dimensions['A'].width = 20
        self.worksheet.column_dimensions['B'].width = 15

        # Save initial file
        self.workbook.save(self.main_file_path)

        # Create initial FINAL file copy
        self.copy_to_final()

        print(f"Created Excel files:")
        print(f"  Main file: {self.main_file_path}")
        print(f"  FINAL file: {self.final_file_path}")
        print(f"  Save interval: every {self.save_interval} measurements")
        print(f"  FINAL update: every 5 minutes")

    def read_vrms(self) -> Optional[float]:
        """
        Read Vrms from Channel 1.

        Returns:
            Vrms value in volts, or None if read failed
        """
        try:
            vrms = self.scope.measure_vrms(channel=1)
            return vrms
        except Exception as e:
            print(f"Error reading Vrms: {e}")
            return None

    def buffer_data(self, timestamp_str: str, vrms: float) -> None:
        """
        Add data to buffer. Writes to file when buffer reaches save_interval.

        Args:
            timestamp_str: Timestamp string
            vrms: Vrms value in volts
        """
        with self.data_lock:
            # Add to buffer
            self.data_buffer.append((timestamp_str, vrms))
            self.measurement_count += 1

            # Check if we need to flush buffer to disk
            if len(self.data_buffer) >= self.save_interval:
                self.flush_buffer()

    def flush_buffer(self) -> None:
        """Write buffered data to Excel file. Must be called with data_lock held."""
        if not self.data_buffer:
            return

        try:
            # Write all buffered data to worksheet
            for timestamp_str, vrms in self.data_buffer:
                self.worksheet.cell(row=self.row_index, column=1, value=timestamp_str)
                self.worksheet.cell(row=self.row_index, column=2, value=vrms)
                self.row_index += 1

            # Save to file once
            self.workbook.save(self.main_file_path)

            # Clear buffer
            self.data_buffer.clear()

        except Exception as e:
            print(f"Error flushing buffer: {e}")

    def copy_to_final(self) -> None:
        """Copy the main file to the FINAL file."""
        try:
            with self.data_lock:
                # Flush any pending data first
                self.flush_buffer()

                # Save current workbook
                self.workbook.save(self.main_file_path)

            # Copy to FINAL file by loading and saving
            # This ensures the FINAL file is not locked by Excel
            wb_copy = load_workbook(self.main_file_path)
            wb_copy.save(self.final_file_path)
            wb_copy.close()

            self.last_copy_time = time.time()
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Updated FINAL file ({self.measurement_count} measurements)")

        except Exception as e:
            print(f"Error copying to FINAL file: {e}")

    def run(self) -> None:
        """Start the data logging loop with precise 100ms timing."""
        print("\n" + "="*60)
        print("Starting real-time Vrms data logging")
        print("="*60)
        print(f"Channel: 1")
        print(f"Sampling interval: 100 ms (precise timing)")
        print(f"FINAL file update interval: 5 minutes")
        print(f"Press Ctrl+C to stop logging")
        print("="*60 + "\n")

        self.running = True
        self.last_copy_time = time.time()

        # Target interval in seconds
        target_interval = 0.1  # 100ms

        # Get initial time
        next_measurement_time = time.time()

        try:
            while self.running:
                # Record current time with milliseconds
                now = datetime.now()
                timestamp_str = now.strftime("%H:%M:%S") + f":{now.microsecond // 1000:03d}"

                # Read Vrms
                vrms = self.read_vrms()

                if vrms is not None:
                    # Buffer data (will auto-save when buffer is full)
                    self.buffer_data(timestamp_str, vrms)

                    # Print to console
                    print(f"{timestamp_str} | {vrms:.6f} V")

                # Check if 5 minutes have passed
                if time.time() - self.last_copy_time >= 300:  # 300 seconds = 5 minutes
                    self.copy_to_final()

                # Calculate next measurement time for precise 100ms intervals
                next_measurement_time += target_interval

                # Calculate how long to sleep
                sleep_time = next_measurement_time - time.time()

                # If we're running behind, skip sleep but adjust next time
                if sleep_time > 0:
                    time.sleep(sleep_time)
                else:
                    # We're running behind schedule
                    # Reset to current time to avoid rapid catch-up
                    if sleep_time < -target_interval:
                        print(f"Warning: Running {-sleep_time:.3f}s behind schedule")
                        next_measurement_time = time.time()

        except KeyboardInterrupt:
            print("\n\nStopping data logging...")
            self.running = False

        finally:
            # Flush any remaining buffered data
            print("Flushing buffered data...")
            with self.data_lock:
                self.flush_buffer()

            # Final copy to FINAL file before exiting
            print("Saving final data...")
            self.copy_to_final()
            print("Data logging completed!")
            print(f"Total measurements: {self.measurement_count}")

    def close(self) -> None:
        """Close all resources."""
        if self.workbook:
            with self.data_lock:
                try:
                    self.flush_buffer()
                    self.workbook.save(self.main_file_path)
                    self.workbook.close()
                except Exception as e:
                    print(f"Error closing workbook: {e}")


def signal_handler(signum, frame):
    """Handle Ctrl+C signal."""
    print("\n\nReceived interrupt signal. Stopping...")
    sys.exit(0)


def main():
    """Main entry point."""
    # Check command line arguments
    if len(sys.argv) < 2:
        print("Error: VISA resource string required")
        print("\nUsage:")
        print("  python vrms_logger.py <RESOURCE_STRING> [save_interval]")
        print("\nExample:")
        print('  python vrms_logger.py "USB0::0x0957::0x17A6::MY12345678::INSTR"')
        print('  python vrms_logger.py "TCPIP0::192.168.1.100::INSTR" 100')
        print("\nOptional:")
        print("  save_interval: Number of measurements before saving (default: 50)")
        sys.exit(1)

    resource_string = sys.argv[1]

    # Optional save interval
    save_interval = 50
    if len(sys.argv) >= 3:
        try:
            save_interval = int(sys.argv[2])
        except ValueError:
            print(f"Warning: Invalid save_interval '{sys.argv[2]}', using default: 50")

    # Set up signal handler for graceful shutdown
    signal.signal(signal.SIGINT, signal_handler)

    # Create logger instance
    logger = VrmsLogger(resource_string, save_interval=save_interval)

    try:
        # Connect to oscilloscope
        logger.connect()

        # Setup Excel files
        logger.setup_excel_files()

        # Start logging
        logger.run()

    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()

    finally:
        # Clean up
        logger.close()
        logger.disconnect()


if __name__ == "__main__":
    main()
