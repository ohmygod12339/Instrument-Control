"""
Fast Real-Time Vrms + Temperature Data Logger
Keysight DSOX4034A (Vrms) + Anbai AT4516 (Temperature)

This script combines oscilloscope Vrms measurements with multi-channel temperature readings.

Features:
- Oscilloscope: Fast Vrms measurement using built-in scope function
- Temperature: AT4516 4-channel thermocouple readings (TC-K type)
- Sampling: Every 1 second
- Buffered writes for performance
- Records Elapsed Time in ms and hr for plotting

Usage:
    python PAPABIN_dsox4034a-at4516_vrms-fast-temp.py [SCOPE_RESOURCE] [TEMP_PORT] [save_interval]

Arguments:
    SCOPE_RESOURCE: VISA resource string for oscilloscope (default: TCPIP::192.168.2.73::INSTR)
    TEMP_PORT: COM port for AT4516 (default: COM10)
    save_interval: Number of measurements before saving to disk (default: 10)

Example:
    # Use all defaults
    python PAPABIN_dsox4034a-at4516_vrms-fast-temp.py

    # Custom resource strings
    python PAPABIN_dsox4034a-at4516_vrms-fast-temp.py "TCPIP::192.168.2.60::INSTR" "COM5"

    # All custom parameters
    python PAPABIN_dsox4034a-at4516_vrms-fast-temp.py "TCPIP::192.168.2.60::INSTR" "COM10" 20
"""

import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Tuple
import threading
import signal

from openpyxl import Workbook, load_workbook
from instruments.keysight.dsox4034a import DSOX4034A
from instruments.anbai import AT4516


class VrmsTempLogger:
    """Combined Vrms + Temperature data logger."""

    def __init__(self, scope_resource: str, temp_port: str, results_dir: str = "results",
                 save_interval: int = 10):
        """
        Initialize the combined logger.

        Args:
            scope_resource: VISA resource string for oscilloscope
            temp_port: COM port for AT4516 temperature meter
            results_dir: Directory to save result files (default: "results")
            save_interval: Number of measurements before saving to disk (default: 10)
        """
        self.scope_resource = scope_resource
        self.temp_port = temp_port
        self.results_dir = Path(results_dir)
        self.save_interval = save_interval

        self.scope = None
        self.temp_meter = None
        self.workbook = None
        self.worksheet = None
        self.main_file_path = None
        self.final_file_path = None
        self.running = False
        self.data_lock = threading.Lock()
        self.row_index = 2
        self.start_time = None
        self.last_copy_time = None
        self.measurement_count = 0

        # Buffer: (timestamp, vrms, temp1, temp2, temp3, temp4, elapsed_ms, elapsed_hr)
        self.data_buffer: List[Tuple[str, float, float, float, float, float, float, float]] = []

        self.results_dir.mkdir(exist_ok=True)

    def connect(self) -> None:
        """Connect to both oscilloscope and temperature meter."""
        print("="*70)
        print("CONNECTING TO INSTRUMENTS")
        print("="*70)

        # Connect to oscilloscope
        print(f"\n1. Oscilloscope: {self.scope_resource}")
        self.scope = DSOX4034A(self.scope_resource)
        self.scope.connect()

        # Configure oscilloscope for fast Vrms measurements
        self.scope.channel_on(1)
        self.scope.set_timebase_scale(0.005)  # 5ms/div for fast measurements
        self.scope.run()

        # Set trigger to AUTO mode for continuous triggering
        self.scope.set_trigger_sweep("AUTO")
        self.scope.set_trigger_level(0.0, 1)
        self.scope.set_trigger_holdoff(0.005)  # 5ms holdoff

        # Set up VRMS measurement on scope
        self.scope.write(":MEAS:CLE")
        self.scope.write(":MEAS:VRMS CHAN1")
        self.scope.write(":MEAS:STAT OFF")
        time.sleep(0.5)

        # Test oscilloscope reading
        try:
            test_vrms = self.read_vrms()
            print(f"   ✓ Oscilloscope ready - Test: {test_vrms:.6f} V")
        except Exception as e:
            print(f"   ✗ Warning: Test measurement failed: {e}")

        # Connect to temperature meter
        print(f"\n2. Temperature Meter: {self.temp_port}")
        self.temp_meter = AT4516(port=self.temp_port)
        self.temp_meter.connect()

        # Configure temperature meter (K-type thermocouples)
        self.temp_meter.configure_and_start(tc_type='TC-K', rate='FAST', unit='CEL')

        # Test temperature reading
        try:
            test_temps = self.read_temperatures()
            print(f"   ✓ Temperature meter ready")
            print(f"   Test readings: Ch1={test_temps[0]:.1f}°C, Ch2={test_temps[1]:.1f}°C, "
                  f"Ch3={test_temps[2]:.1f}°C, Ch4={test_temps[3]:.1f}°C")
        except Exception as e:
            print(f"   ✗ Warning: Test reading failed: {e}")

        print("\n" + "="*70)
        print("✓ ALL INSTRUMENTS CONNECTED AND CONFIGURED")
        print("="*70 + "\n")

    def disconnect(self) -> None:
        """Disconnect from all instruments."""
        if self.scope:
            self.scope.disconnect()
            print("Disconnected from oscilloscope")
        if self.temp_meter:
            self.temp_meter.disconnect()
            print("Disconnected from temperature meter")

    def setup_excel_files(self) -> None:
        """Create and initialize Excel files for data logging."""
        self.start_time = datetime.now()
        timestamp = self.start_time.strftime("%Y%m%d_%H%M%S")

        self.main_file_path = self.results_dir / f"Result_{timestamp}.xlsx"
        self.final_file_path = self.results_dir / f"Result_{timestamp}_FINAL.xlsx"

        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Vrms + Temperature"

        # Headers: Timestamp, Vrms, Temp1-4, Elapsed Time (ms), Elapsed Time (hr)
        headers = [
            "Timestamp",
            "Vrms (V)",
            "Temp Ch1 (°C)",
            "Temp Ch2 (°C)",
            "Temp Ch3 (°C)",
            "Temp Ch4 (°C)",
            "Elapsed Time (ms)",
            "Elapsed Time (hr)"
        ]
        self.worksheet.append(headers)

        # Set column widths
        self.worksheet.column_dimensions['A'].width = 20  # Timestamp
        self.worksheet.column_dimensions['B'].width = 15  # Vrms
        self.worksheet.column_dimensions['C'].width = 15  # Temp Ch1
        self.worksheet.column_dimensions['D'].width = 15  # Temp Ch2
        self.worksheet.column_dimensions['E'].width = 15  # Temp Ch3
        self.worksheet.column_dimensions['F'].width = 15  # Temp Ch4
        self.worksheet.column_dimensions['G'].width = 20  # Elapsed Time (ms)
        self.worksheet.column_dimensions['H'].width = 20  # Elapsed Time (hr)

        self.workbook.save(self.main_file_path)
        self.copy_to_final()

        print(f"Created Excel files:")
        print(f"  Main file: {self.main_file_path}")
        print(f"  FINAL file: {self.final_file_path}")
        print(f"  Save interval: every {self.save_interval} measurements")

    def read_vrms(self) -> Optional[float]:
        """
        Read Vrms from oscilloscope.

        Returns:
            Vrms value in volts, or None if read failed
        """
        try:
            vrms_str = self.scope.query(":MEAS:VRMS? CHAN1")
            vrms = float(vrms_str)
            return vrms
        except Exception as e:
            print(f"Error reading Vrms: {e}")
            return None

    def read_temperatures(self) -> List[Optional[float]]:
        """
        Read temperatures from AT4516 channels 1-4.

        Returns:
            List of 4 temperature values (°C), or None for channels that failed
        """
        try:
            all_temps = self.temp_meter.read_all_channels()
            # Get first 4 channels, handle if less than 4 returned
            temps = []
            for i in range(4):
                if i < len(all_temps):
                    temp = all_temps[i]
                    # Filter out error values (-100000.0)
                    if temp is not None and temp > -100000:
                        temps.append(temp)
                    else:
                        temps.append(None)
                else:
                    temps.append(None)
            return temps
        except Exception as e:
            print(f"Error reading temperatures: {e}")
            return [None, None, None, None]

    def buffer_data(self, timestamp_str: str, vrms: float, temps: List[float], elapsed_ms: float) -> None:
        """Add data to buffer and flush when full."""
        with self.data_lock:
            elapsed_hr = elapsed_ms / 3_600_000  # Convert ms to hours

            # Unpack temperature values
            temp1, temp2, temp3, temp4 = temps[0], temps[1], temps[2], temps[3]

            self.data_buffer.append((timestamp_str, vrms, temp1, temp2, temp3, temp4, elapsed_ms, elapsed_hr))
            self.measurement_count += 1

            if len(self.data_buffer) >= self.save_interval:
                self.flush_buffer()

    def flush_buffer(self) -> None:
        """Write buffered data to Excel file."""
        if not self.data_buffer:
            return

        try:
            for timestamp_str, vrms, temp1, temp2, temp3, temp4, elapsed_ms, elapsed_hr in self.data_buffer:
                self.worksheet.cell(row=self.row_index, column=1, value=timestamp_str)
                self.worksheet.cell(row=self.row_index, column=2, value=vrms)
                self.worksheet.cell(row=self.row_index, column=3, value=temp1)
                self.worksheet.cell(row=self.row_index, column=4, value=temp2)
                self.worksheet.cell(row=self.row_index, column=5, value=temp3)
                self.worksheet.cell(row=self.row_index, column=6, value=temp4)
                self.worksheet.cell(row=self.row_index, column=7, value=elapsed_ms)
                self.worksheet.cell(row=self.row_index, column=8, value=elapsed_hr)
                self.row_index += 1

            self.workbook.save(self.main_file_path)
            self.data_buffer.clear()

        except Exception as e:
            print(f"Error flushing buffer: {e}")

    def copy_to_final(self) -> None:
        """Copy the main file to the FINAL file."""
        try:
            with self.data_lock:
                self.flush_buffer()
                self.workbook.save(self.main_file_path)

            wb_copy = load_workbook(self.main_file_path)
            wb_copy.save(self.final_file_path)
            wb_copy.close()

            self.last_copy_time = time.time()
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Updated FINAL file ({self.measurement_count} measurements)")

        except Exception as e:
            print(f"Error copying to FINAL file: {e}")

    def run(self) -> None:
        """Start the data logging loop with 1-second intervals."""
        print("\n" + "="*70)
        print("STARTING DATA LOGGING")
        print("="*70)
        print(f"Oscilloscope: Channel 1 Vrms")
        print(f"Temperature: Channels 1-4 (K-type thermocouples)")
        print(f"Sampling interval: 1.0 second")
        print(f"FINAL file update: 5 minutes")
        print(f"Press Ctrl+C to stop logging")
        print("="*70)
        print(f"\n{'Timestamp':<15} {'Vrms (V)':<12} {'T1(°C)':<10} {'T2(°C)':<10} {'T3(°C)':<10} {'T4(°C)':<10} {'Elapsed(ms)':<15}")
        print("-"*90)

        self.running = True
        self.last_copy_time = time.time()

        target_interval = 1.0  # 1 second
        next_measurement_time = time.time()

        try:
            while self.running:
                now = datetime.now()
                timestamp_str = now.strftime("%H:%M:%S") + f":{now.microsecond // 1000:03d}"

                # Calculate elapsed time in milliseconds from start
                elapsed_ms = (now - self.start_time).total_seconds() * 1000

                # Read Vrms from oscilloscope
                vrms = self.read_vrms()

                # Read temperatures from AT4516
                temps = self.read_temperatures()

                if vrms is not None:
                    self.buffer_data(timestamp_str, vrms, temps, elapsed_ms)

                    # Format temperature values for display
                    temp_strs = []
                    for temp in temps:
                        if temp is not None:
                            temp_strs.append(f"{temp:7.2f}")
                        else:
                            temp_strs.append("   N/A ")

                    print(f"{timestamp_str:<15} {vrms:10.6f}  "
                          f"{temp_strs[0]:<10} {temp_strs[1]:<10} {temp_strs[2]:<10} {temp_strs[3]:<10} "
                          f"{elapsed_ms:13.1f}")

                # Check if 5 minutes have passed for FINAL file update
                if time.time() - self.last_copy_time >= 300:
                    self.copy_to_final()

                # Precise timing - next measurement in 1 second
                next_measurement_time += target_interval
                sleep_time = next_measurement_time - time.time()

                if sleep_time > 0:
                    time.sleep(sleep_time)
                else:
                    if sleep_time < -target_interval:
                        print(f"Warning: Running {-sleep_time:.3f}s behind schedule")
                        next_measurement_time = time.time()

        except KeyboardInterrupt:
            print("\n\nStopping data logging...")
            self.running = False

        finally:
            print("Flushing buffered data...")
            with self.data_lock:
                self.flush_buffer()

            print("Saving final data...")
            self.copy_to_final()
            print("Data logging completed!")
            print(f"Total measurements: {self.measurement_count}")

    def close(self) -> None:
        """Close all resources."""
        if self.workbook:
            with self.data_lock:
                try:
                    self.flush_buffer()
                    self.workbook.save(self.main_file_path)
                    self.workbook.close()
                except Exception as e:
                    print(f"Error closing workbook: {e}")


def signal_handler(signum, frame):
    """Handle Ctrl+C signal."""
    print("\n\nReceived interrupt signal. Stopping...")
    sys.exit(0)


def main():
    """Main entry point."""
    # Default values
    default_scope_resource = "TCPIP::192.168.2.73::INSTR"
    default_temp_port = "COM10"
    default_save_interval = 10

    # Show help if requested
    if len(sys.argv) >= 2 and sys.argv[1] in ['-h', '--help']:
        print("Usage:")
        print("  python PAPABIN_dsox4034a-at4516_vrms-fast-temp.py [SCOPE_RESOURCE] [TEMP_PORT] [save_interval]")
        print("\nArguments:")
        print(f"  SCOPE_RESOURCE: VISA resource for oscilloscope (default: {default_scope_resource})")
        print(f"  TEMP_PORT: COM port for AT4516 (default: {default_temp_port})")
        print(f"  save_interval: Measurements before saving (default: {default_save_interval})")
        print("\nExamples:")
        print('  python PAPABIN_dsox4034a-at4516_vrms-fast-temp.py')
        print('  python PAPABIN_dsox4034a-at4516_vrms-fast-temp.py "TCPIP::192.168.2.60::INSTR" "COM5"')
        print('  python PAPABIN_dsox4034a-at4516_vrms-fast-temp.py "TCPIP::192.168.2.60::INSTR" "COM10" 20')
        sys.exit(0)

    # Parse arguments
    scope_resource = default_scope_resource
    if len(sys.argv) >= 2:
        scope_resource = sys.argv[1]

    temp_port = default_temp_port
    if len(sys.argv) >= 3:
        temp_port = sys.argv[2]

    save_interval = default_save_interval
    if len(sys.argv) >= 4:
        try:
            save_interval = int(sys.argv[3])
        except ValueError:
            print(f"Warning: Invalid save_interval '{sys.argv[3]}', using default: {default_save_interval}")

    # Print configuration
    print(f"Configuration:")
    print(f"  Oscilloscope: {scope_resource}")
    print(f"  Temperature: {temp_port}")
    print(f"  Save interval: {save_interval}")

    signal.signal(signal.SIGINT, signal_handler)

    logger = VrmsTempLogger(scope_resource, temp_port, save_interval=save_interval)

    try:
        logger.connect()
        logger.setup_excel_files()
        logger.run()

    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()

    finally:
        logger.close()
        logger.disconnect()


if __name__ == "__main__":
    main()
