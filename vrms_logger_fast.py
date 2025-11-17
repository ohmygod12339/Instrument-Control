"""
Fast Real-Time Vrms Data Logger for Keysight DSOX4034A

This version uses the oscilloscope's built-in measurement feature for faster acquisition.
Instead of triggering a new measurement each time, it reads the continuously-updated
measurement value from the scope's display.

Features:
- Uses scope's built-in VRMS measurement (faster)
- Reads continuously-updated measurement value
- No trigger wait - reads current displayed value
- Configurable timebase for optimal measurement speed
- Buffered writes for performance
- Precise 100ms timing

Usage:
    python vrms_logger_fast.py <RESOURCE_STRING> [save_interval] [timebase_ms]

Arguments:
    RESOURCE_STRING: VISA resource string for the oscilloscope (required)
    save_interval: Number of measurements before saving to disk (default: 50)
    timebase_ms: Oscilloscope timebase in milliseconds/div (default: 10)
                 Shorter timebase = faster measurements
                 Recommended: 5-20 ms/div for AC signals

Example:
    # Use default settings (10 ms/div timebase)
    python vrms_logger_fast.py "TCPIP::192.168.2.60::INSTR"

    # Custom save interval and 5 ms/div timebase for fastest measurements
    python vrms_logger_fast.py "TCPIP::192.168.2.60::INSTR" 100 5

    # 20 ms/div timebase for more stable measurements
    python vrms_logger_fast.py "TCPIP::192.168.2.60::INSTR" 50 20
"""

import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Tuple
import threading
import signal

from openpyxl import Workbook, load_workbook
from instruments.keysight.dsox4034a import DSOX4034A


class FastVrmsLogger:
    """Fast real-time Vrms data logger using scope's built-in measurements."""

    def __init__(self, resource_string: str, results_dir: str = "results",
                 save_interval: int = 50, timebase_scale: float = 0.01):
        """
        Initialize the fast Vrms logger.

        Args:
            resource_string: VISA resource string for the oscilloscope
            results_dir: Directory to save result files (default: "results")
            save_interval: Number of measurements before saving to disk (default: 50)
            timebase_scale: Oscilloscope timebase in seconds/div (default: 0.01 = 10ms/div)
                           Shorter timebase = faster measurements
                           Recommended: 0.005 (5ms/div) to 0.02 (20ms/div) for AC signals
        """
        self.resource_string = resource_string
        self.results_dir = Path(results_dir)
        self.save_interval = save_interval
        self.timebase_scale = timebase_scale
        self.scope = None
        self.workbook = None
        self.worksheet = None
        self.main_file_path = None
        self.final_file_path = None
        self.running = False
        self.data_lock = threading.Lock()
        self.row_index = 2
        self.start_time = None
        self.last_copy_time = None
        self.measurement_count = 0
        self.data_buffer: List[Tuple[str, float]] = []

        self.results_dir.mkdir(exist_ok=True)

    def connect(self) -> None:
        """Connect to the oscilloscope and configure for fast measurements."""
        print(f"Connecting to oscilloscope: {self.resource_string}")
        self.scope = DSOX4034A(self.resource_string)
        self.scope.connect()

        # Ensure Channel 1 is on
        self.scope.channel_on(1)

        # CRITICAL: Set timebase for fast measurements
        # Shorter timebase = less data to capture = faster measurements
        print(f"Setting timebase to {self.timebase_scale*1000:.1f} ms/div...")
        self.scope.set_timebase_scale(self.timebase_scale)
        current_timebase = self.scope.get_timebase_scale()
        print(f"  Timebase confirmed: {current_timebase*1000:.1f} ms/div ({current_timebase*10*1000:.1f} ms total window)")

        # Set oscilloscope to RUN mode for continuous acquisition
        self.scope.run()

        # IMPORTANT: Set up measurement on the scope itself
        # This makes the scope continuously calculate VRMS and we just read the result
        print("Configuring scope measurements...")

        # Clear any existing measurements
        self.scope.write(":MEAS:CLE")

        # Add VRMS measurement for Channel 1
        # This makes the scope display and continuously update VRMS
        self.scope.write(":MEAS:VRMS CHAN1")

        # Turn off statistics to speed up (we don't need min/max/avg/stddev)
        self.scope.write(":MEAS:STAT OFF")

        # Wait a moment for measurement to stabilize
        time.sleep(0.5)

        # Test read
        try:
            test_vrms = self.read_vrms_fast()
            print(f"Test measurement: {test_vrms:.6f} V")
        except Exception as e:
            print(f"Warning: Test measurement failed: {e}")

        print(f"Successfully connected and configured!")
        print(f"  Channel 1: ON")
        print(f"  Timebase: {current_timebase*1000:.1f} ms/div")
        print(f"  Mode: RUN (continuous)")
        print(f"  Measurement: VRMS (built-in)")

    def disconnect(self) -> None:
        """Disconnect from the oscilloscope."""
        if self.scope:
            self.scope.disconnect()
            print("Disconnected from oscilloscope")

    def setup_excel_files(self) -> None:
        """Create and initialize Excel files for data logging."""
        self.start_time = datetime.now()
        timestamp = self.start_time.strftime("%Y%m%d_%H%M%S")

        self.main_file_path = self.results_dir / f"Result_{timestamp}_Real-Time-Result.xlsx"
        self.final_file_path = self.results_dir / f"Result_{timestamp}_Real-Time-Result_FINAL.xlsx"

        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Vrms Data"

        self.worksheet.append(["Timestamp", "Vrms (V)"])
        self.worksheet.column_dimensions['A'].width = 20
        self.worksheet.column_dimensions['B'].width = 15

        self.workbook.save(self.main_file_path)
        self.copy_to_final()

        print(f"Created Excel files:")
        print(f"  Main file: {self.main_file_path}")
        print(f"  FINAL file: {self.final_file_path}")
        print(f"  Save interval: every {self.save_interval} measurements")

    def read_vrms_fast(self) -> Optional[float]:
        """
        Read the continuously-updated VRMS value from scope's display.
        This is much faster than triggering a new measurement each time.

        Returns:
            Vrms value in volts, or None if read failed
        """
        try:
            # Query the displayed VRMS result (should be fast)
            # This reads the current value without triggering a new acquisition
            vrms_str = self.scope.query(":MEAS:VRMS? CHAN1")
            vrms = float(vrms_str)
            return vrms
        except Exception as e:
            print(f"Error reading Vrms: {e}")
            return None

    def buffer_data(self, timestamp_str: str, vrms: float) -> None:
        """Add data to buffer and flush when full."""
        with self.data_lock:
            self.data_buffer.append((timestamp_str, vrms))
            self.measurement_count += 1

            if len(self.data_buffer) >= self.save_interval:
                self.flush_buffer()

    def flush_buffer(self) -> None:
        """Write buffered data to Excel file."""
        if not self.data_buffer:
            return

        try:
            for timestamp_str, vrms in self.data_buffer:
                self.worksheet.cell(row=self.row_index, column=1, value=timestamp_str)
                self.worksheet.cell(row=self.row_index, column=2, value=vrms)
                self.row_index += 1

            self.workbook.save(self.main_file_path)
            self.data_buffer.clear()

        except Exception as e:
            print(f"Error flushing buffer: {e}")

    def copy_to_final(self) -> None:
        """Copy the main file to the FINAL file."""
        try:
            with self.data_lock:
                self.flush_buffer()
                self.workbook.save(self.main_file_path)

            wb_copy = load_workbook(self.main_file_path)
            wb_copy.save(self.final_file_path)
            wb_copy.close()

            self.last_copy_time = time.time()
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Updated FINAL file ({self.measurement_count} measurements)")

        except Exception as e:
            print(f"Error copying to FINAL file: {e}")

    def run(self) -> None:
        """Start the data logging loop with precise 100ms timing."""
        print("\n" + "="*60)
        print("Starting FAST real-time Vrms data logging")
        print("="*60)
        print(f"Channel: 1")
        print(f"Sampling interval: 100 ms (precise timing)")
        print(f"Method: Built-in scope measurement (fast)")
        print(f"FINAL file update: 5 minutes")
        print(f"Press Ctrl+C to stop logging")
        print("="*60 + "\n")

        self.running = True
        self.last_copy_time = time.time()

        target_interval = 0.1  # 100ms
        next_measurement_time = time.time()

        try:
            while self.running:
                now = datetime.now()
                timestamp_str = now.strftime("%H:%M:%S") + f":{now.microsecond // 1000:03d}"

                # Read the continuously-updated VRMS (should be fast!)
                vrms = self.read_vrms_fast()

                if vrms is not None:
                    self.buffer_data(timestamp_str, vrms)
                    print(f"{timestamp_str} | {vrms:.6f} V")

                # Check if 5 minutes have passed
                if time.time() - self.last_copy_time >= 300:
                    self.copy_to_final()

                # Precise timing
                next_measurement_time += target_interval
                sleep_time = next_measurement_time - time.time()

                if sleep_time > 0:
                    time.sleep(sleep_time)
                else:
                    if sleep_time < -target_interval:
                        print(f"Warning: Running {-sleep_time:.3f}s behind schedule")
                        next_measurement_time = time.time()

        except KeyboardInterrupt:
            print("\n\nStopping data logging...")
            self.running = False

        finally:
            print("Flushing buffered data...")
            with self.data_lock:
                self.flush_buffer()

            print("Saving final data...")
            self.copy_to_final()
            print("Data logging completed!")
            print(f"Total measurements: {self.measurement_count}")

    def close(self) -> None:
        """Close all resources."""
        if self.workbook:
            with self.data_lock:
                try:
                    self.flush_buffer()
                    self.workbook.save(self.main_file_path)
                    self.workbook.close()
                except Exception as e:
                    print(f"Error closing workbook: {e}")


def signal_handler(signum, frame):
    """Handle Ctrl+C signal."""
    print("\n\nReceived interrupt signal. Stopping...")
    sys.exit(0)


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print("Error: VISA resource string required")
        print("\nUsage:")
        print("  python vrms_logger_fast.py <RESOURCE_STRING> [save_interval] [timebase_ms]")
        print("\nArguments:")
        print("  RESOURCE_STRING: VISA resource string (required)")
        print("  save_interval: Number of measurements before saving (default: 50)")
        print("  timebase_ms: Timebase in ms/div (default: 10)")
        print("               Shorter = faster, Recommended: 5-20 ms/div")
        print("\nExamples:")
        print('  python vrms_logger_fast.py "TCPIP::192.168.2.60::INSTR"')
        print('  python vrms_logger_fast.py "TCPIP::192.168.2.60::INSTR" 100 5')
        sys.exit(1)

    resource_string = sys.argv[1]

    # Parse save_interval (argument 2)
    save_interval = 50
    if len(sys.argv) >= 3:
        try:
            save_interval = int(sys.argv[2])
        except ValueError:
            print(f"Warning: Invalid save_interval '{sys.argv[2]}', using default: 50")

    # Parse timebase_ms (argument 3)
    timebase_scale = 0.01  # Default: 10 ms/div = 0.01 s/div
    if len(sys.argv) >= 4:
        try:
            timebase_ms = float(sys.argv[3])
            timebase_scale = timebase_ms / 1000.0  # Convert ms to seconds
            print(f"Using timebase: {timebase_ms} ms/div ({timebase_scale} s/div)")
        except ValueError:
            print(f"Warning: Invalid timebase_ms '{sys.argv[3]}', using default: 10 ms/div")

    signal.signal(signal.SIGINT, signal_handler)

    logger = FastVrmsLogger(resource_string, save_interval=save_interval,
                           timebase_scale=timebase_scale)

    try:
        logger.connect()
        logger.setup_excel_files()
        logger.run()

    except Exception as e:
        print(f"\nError: {e}")
        import traceback
        traceback.print_exc()

    finally:
        logger.close()
        logger.disconnect()


if __name__ == "__main__":
    main()
