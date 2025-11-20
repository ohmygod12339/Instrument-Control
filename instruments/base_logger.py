"""
Base Data Logger Template

This module provides a base class for real-time data logging with:
- Precise timing compensation (not affected by processing time)
- Buffered Excel output for performance
- Elapsed time tracking
- Dual-file system (main + FINAL for viewing)
- Graceful shutdown handling

Inherit from this class to create instrument-specific loggers.

Example:
    class MyLogger(BaseDataLogger):
        def setup_instrument(self):
            self.instrument = MyInstrument()
            self.instrument.connect()

        def read_measurement(self):
            return self.instrument.read_value()

        def cleanup_instrument(self):
            self.instrument.disconnect()

        def get_headers(self):
            return ["Timestamp", "Value", "Elapsed Time (ms)", "Elapsed Time (hr)"]
"""

import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Any
from abc import ABC, abstractmethod
import threading
import signal

from openpyxl import Workbook, load_workbook


class BaseDataLogger(ABC):
    """
    Abstract base class for real-time data logging.

    Provides precise timing, buffered writes, and Excel output.
    Subclasses must implement:
        - setup_instrument(): Initialize and configure the instrument
        - read_measurement(): Read a single measurement value
        - cleanup_instrument(): Disconnect and cleanup
        - get_headers(): Return column headers for Excel
    """

    def __init__(self, results_dir: str = "results", save_interval: int = 50,
                 measurement_interval: float = 0.1, final_copy_interval: float = 300):
        """
        Initialize the data logger.

        Args:
            results_dir: Directory to save result files (default: "results")
            save_interval: Number of measurements before saving to disk (default: 50)
            measurement_interval: Time between measurements in seconds (default: 0.1 = 100ms)
            final_copy_interval: Time between FINAL file updates in seconds (default: 300 = 5min)
        """
        self.results_dir = Path(results_dir)
        self.save_interval = save_interval
        self.measurement_interval = measurement_interval
        self.final_copy_interval = final_copy_interval

        self.workbook = None
        self.worksheet = None
        self.main_file_path = None
        self.final_file_path = None
        self.running = False
        self.data_lock = threading.Lock()
        self.row_index = 2
        self.start_time = None
        self.last_copy_time = None
        self.measurement_count = 0
        self.data_buffer: List[List[Any]] = []

        self.results_dir.mkdir(exist_ok=True)

    @abstractmethod
    def setup_instrument(self) -> None:
        """
        Initialize and configure the instrument.

        This method should:
        - Create the instrument instance
        - Connect to the instrument
        - Configure measurement settings
        """
        pass

    @abstractmethod
    def read_measurement(self) -> Any:
        """
        Read a single measurement from the instrument.

        Returns:
            The measurement value (float, tuple, or any type)
            Return None if measurement failed
        """
        pass

    @abstractmethod
    def cleanup_instrument(self) -> None:
        """
        Disconnect and cleanup the instrument.
        """
        pass

    @abstractmethod
    def get_headers(self) -> List[str]:
        """
        Return the column headers for the Excel file.

        Returns:
            List of header strings, e.g., ["Timestamp", "Value", "Elapsed Time (ms)", "Elapsed Time (hr)"]
        """
        pass

    def format_measurement(self, timestamp_str: str, elapsed_ms: float,
                          measurement: Any) -> List[Any]:
        """
        Format a measurement into a row for the Excel file.

        Override this method if you need custom formatting.
        Default implementation handles single values and tuples.

        Args:
            timestamp_str: Formatted timestamp string
            elapsed_ms: Elapsed time in milliseconds
            measurement: The measurement value(s)

        Returns:
            List of values for the Excel row
        """
        elapsed_hr = elapsed_ms / 3_600_000  # Convert ms to hours
        if isinstance(measurement, (list, tuple)):
            return [timestamp_str, *measurement, elapsed_ms, elapsed_hr]
        else:
            return [timestamp_str, measurement, elapsed_ms, elapsed_hr]

    def format_display(self, timestamp_str: str, elapsed_ms: float,
                      measurement: Any) -> str:
        """
        Format a measurement for console display.

        Override this method for custom display formatting.

        Args:
            timestamp_str: Formatted timestamp string
            elapsed_ms: Elapsed time in milliseconds
            measurement: The measurement value(s)

        Returns:
            Formatted string for console output
        """
        if isinstance(measurement, (list, tuple)):
            values = " | ".join(f"{v:.6f}" if isinstance(v, float) else str(v)
                               for v in measurement)
            return f"{timestamp_str} | {values} | {elapsed_ms:.1f} ms"
        elif isinstance(measurement, float):
            return f"{timestamp_str} | {measurement:.6f} | {elapsed_ms:.1f} ms"
        else:
            return f"{timestamp_str} | {measurement} | {elapsed_ms:.1f} ms"

    def setup_excel_files(self) -> None:
        """Create and initialize Excel files for data logging."""
        self.start_time = datetime.now()
        timestamp = self.start_time.strftime("%Y%m%d_%H%M%S")

        self.main_file_path = self.results_dir / f"Result_{timestamp}.xlsx"
        self.final_file_path = self.results_dir / f"Result_{timestamp}_FINAL.xlsx"

        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Data"

        # Write headers
        headers = self.get_headers()
        self.worksheet.append(headers)

        # Set column widths
        for i, header in enumerate(headers, 1):
            col_letter = chr(64 + i)  # A, B, C, ...
            self.worksheet.column_dimensions[col_letter].width = max(len(header) + 2, 15)

        self.workbook.save(self.main_file_path)
        self.copy_to_final()

        print(f"Created Excel files:")
        print(f"  Main file: {self.main_file_path}")
        print(f"  FINAL file: {self.final_file_path}")
        print(f"  Save interval: every {self.save_interval} measurements")

    def buffer_data(self, row: List[Any]) -> None:
        """Add data to buffer and flush when full."""
        with self.data_lock:
            self.data_buffer.append(row)
            self.measurement_count += 1

            if len(self.data_buffer) >= self.save_interval:
                self.flush_buffer()

    def flush_buffer(self) -> None:
        """Write buffered data to Excel file."""
        if not self.data_buffer:
            return

        try:
            for row in self.data_buffer:
                for col, value in enumerate(row, 1):
                    self.worksheet.cell(row=self.row_index, column=col, value=value)
                self.row_index += 1

            self.workbook.save(self.main_file_path)
            self.data_buffer.clear()

        except Exception as e:
            print(f"Error flushing buffer: {e}")

    def copy_to_final(self) -> None:
        """Copy the main file to the FINAL file."""
        try:
            with self.data_lock:
                self.flush_buffer()
                self.workbook.save(self.main_file_path)

            wb_copy = load_workbook(self.main_file_path)
            wb_copy.save(self.final_file_path)
            wb_copy.close()

            self.last_copy_time = time.time()
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Updated FINAL file ({self.measurement_count} measurements)")

        except Exception as e:
            print(f"Error copying to FINAL file: {e}")

    def run(self) -> None:
        """Start the data logging loop with precise timing."""
        interval_ms = self.measurement_interval * 1000

        print("\n" + "="*60)
        print("Starting real-time data logging")
        print("="*60)
        print(f"Sampling interval: {interval_ms:.0f} ms (precise timing)")
        print(f"FINAL file update: {self.final_copy_interval / 60:.1f} minutes")
        print(f"Press Ctrl+C to stop logging")
        print("="*60 + "\n")

        self.running = True
        self.last_copy_time = time.time()

        # Precise timing variables
        next_measurement_time = time.time()

        try:
            while self.running:
                # Get current time
                now = datetime.now()
                timestamp_str = now.strftime("%H:%M:%S") + f":{now.microsecond // 1000:03d}"

                # Calculate elapsed time in milliseconds from start
                elapsed_ms = (now - self.start_time).total_seconds() * 1000

                # Read measurement
                measurement = self.read_measurement()

                if measurement is not None:
                    # Format and buffer the data
                    row = self.format_measurement(timestamp_str, elapsed_ms, measurement)
                    self.buffer_data(row)

                    # Display to console
                    display = self.format_display(timestamp_str, elapsed_ms, measurement)
                    print(display)

                # Check if it's time to update FINAL file
                if time.time() - self.last_copy_time >= self.final_copy_interval:
                    self.copy_to_final()

                # Precise timing compensation
                next_measurement_time += self.measurement_interval
                sleep_time = next_measurement_time - time.time()

                if sleep_time > 0:
                    time.sleep(sleep_time)
                else:
                    # Running behind schedule
                    if sleep_time < -self.measurement_interval:
                        print(f"Warning: Running {-sleep_time:.3f}s behind schedule")
                        next_measurement_time = time.time()

        except KeyboardInterrupt:
            print("\n\nStopping data logging...")
            self.running = False

        finally:
            print("Flushing buffered data...")
            with self.data_lock:
                self.flush_buffer()

            print("Saving final data...")
            self.copy_to_final()
            print("Data logging completed!")
            print(f"Total measurements: {self.measurement_count}")

    def close(self) -> None:
        """Close all resources."""
        if self.workbook:
            with self.data_lock:
                try:
                    self.flush_buffer()
                    self.workbook.save(self.main_file_path)
                    self.workbook.close()
                except Exception as e:
                    print(f"Error closing workbook: {e}")

    def start(self) -> None:
        """
        Main entry point - setup, run, and cleanup.

        Call this method to start the logger.
        """
        try:
            self.setup_instrument()
            self.setup_excel_files()
            self.run()

        except Exception as e:
            print(f"\nError: {e}")
            import traceback
            traceback.print_exc()

        finally:
            self.close()
            self.cleanup_instrument()
